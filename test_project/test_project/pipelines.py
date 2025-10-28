import os
import pandas as pd
from loguru import logger
from openpyxl import load_workbook

class AppendPipeline:
    ''' 
        Pipeline, built on pd.ExcelWriter
        to keep styles and formating of the existing sheet.
    '''

    def __init__(self):
        self.items = []
        self.filename = os.path.join(os.getcwd(), "Data Entry - Advanced Content Scraper.xlsx")

    def process_item(self, item, spider):
        self.items.append(item)
        return item

    def close_spider(self, spider):
        if not self.items:
            logger.warning("No items to save.")
            return
        
        if not os.path.exists(self.filename):
            logger.error(f"File not found: {self.filename}")
            return

        logger.info("Processing data with Pandas...")
        
        df = pd.DataFrame(self.items)
        
        df['Name'] = df['Name'].str.strip()
        df['Affiliation'] = df['Affiliation'].str.strip()
        logger.info(f"Cleaned {len(df)} rows using Pandas")
        
        initial_count = len(df)
        df = df.drop_duplicates(subset=['Name', 'Presentation Number'], keep='first')
        duplicates_removed = initial_count - len(df)
        if duplicates_removed > 0:
            logger.info(f"Removed {duplicates_removed} duplicate entries using Pandas")
        
        df = df.sort_values(by=['Session Name', 'Presentation Number', 'Name'])
        
        df = df.rename(columns={
            'Name': 'Name (incl. titles if any mentioned)',
            'Affiliation': 'Affiliation(s)',
            'Role': "Person's role"
        })

        try:
            book = load_workbook(self.filename)
            sheet_name = book.sheetnames[0]
            start_row = 2

            with pd.ExcelWriter(
                self.filename,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="overlay"
            ) as writer:
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=False,
                    startrow=start_row
                )

            logger.info(f"Successfully saved {len(df)} processed rows to {self.filename}")

        except Exception as e:
            logger.error(f"Failed to write data: {e}")


class FilterAbstractAuthorPipeline:
    ''' 
        Pipeline, built with pd.to_excel.
        Create a new file without any predetermined styles and formating.
    '''
        
    def __init__(self):
        self.items = []
        self.filename = os.path.join(os.getcwd(), "Data Entry - Filtered Authors.xlsx")

    def process_item(self, item, spider):
        self.items.append(item)
        return item

    def close_spider(self, spider):
        def parse_name(name):
            parts = name.split()
            if len(parts) == 1:
                return pd.Series({
                    'First Name': parts[0], 
                    'Middle Name': None, 
                    'Last Name': None
                })
            elif len(parts) == 2:
                return pd.Series({
                    'First Name': parts[0], 
                    'Middle Name': None, 
                    'Last Name': parts[1]
                })
            else:
                return pd.Series({
                    'First Name': parts[0], 
                    'Middle Name': ' '.join(parts[1:-1]), 
                    'Last Name': parts[-1]
                })
    
        if not self.items:
            logger.warning("No Abstract authors to save.")
            return
        
        df = pd.DataFrame(self.items)
        
        df_filtered = df[df['Role'] == 'Abstract author'].copy()
        if len(df_filtered) == 0:
            logger.warning("No Abstract authors found after filtering")
            return
        
        logger.info("Parsing names into First, Middle, Last using Pandas apply()")
        name_parts = df_filtered['Name'].apply(parse_name)
        
        df_final = pd.concat([
            name_parts,
            df_filtered[['Affiliation', 'Session Name', 'Presentation Number', 
                        'Topic Title', 'Presentation Abstract', 'Abstract URL']]
        ], axis=1)
        
        df_final = df_final.rename(columns={'Affiliation': 'Affiliation(s)'})
        
        df_final = df_final.sort_values(by=['Session Name', 'Presentation Number', 'Last Name'])

        try:
            df_final.to_excel(
                self.filename, 
                index=False, 
                engine='openpyxl'
            )
            logger.info(f"Successfully saved {len(df_final)} filtered and processed Abstract authors to {self.filename}")
        except Exception as e:
            logger.error(f"Failed to save file with Pandas: {e}")